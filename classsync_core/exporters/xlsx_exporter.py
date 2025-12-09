"""
XLSX (Excel) exporter with styling and formatting.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, Optional
import os

from classsync_core.exports import BaseExporter


class XLSXExporter(BaseExporter):
    """Export timetables to Excel format with styling."""

    def __init__(self, db):
        super().__init__(db)

        # Default styling
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.day_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, timetable_id: int, output_path: str, **kwargs) -> str:
        """
        Export timetable to Excel file.

        Args:
            timetable_id: ID of timetable to export
            output_path: Path where file should be saved
            **kwargs: Options like 'view_type' (section/teacher/room/master)

        Returns:
            Path to exported file
        """
        view_type = kwargs.get('view_type', 'master')

        # Load data
        df = self.load_timetable_data(timetable_id)

        if df.empty:
            raise ValueError(f"No data found for timetable {timetable_id}")

        # Create output directory if needed
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

        if view_type == 'master':
            return self._export_master(df, output_path)
        elif view_type == 'section':
            return self._export_by_section(df, output_path)
        elif view_type == 'teacher':
            return self._export_by_teacher(df, output_path)
        elif view_type == 'room':
            return self._export_by_room(df, output_path)
        else:
            raise ValueError(f"Unknown view type: {view_type}")

    def _export_master(self, df: pd.DataFrame, output_path: str) -> str:
        """Export complete timetable as single Excel file."""

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Timetable"

        # Group by day for better visualization
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        row = 1
        for day in days:
            day_df = df[df['Weekday'] == day].sort_values('Start_Time')

            if day_df.empty:
                continue

            # Day header
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = day.upper()
            cell.fill = self.day_fill
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Column headers
            headers = ['Time', 'Course', 'Section', 'Instructor', 'Room', 'Building', 'Type', 'Duration']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            row += 1

            # Data rows
            for _, entry in day_df.iterrows():
                ws.cell(row=row, column=1, value=f"{entry['Start_Time']} - {entry['End_Time']}")
                ws.cell(row=row, column=2, value=entry['Course_Name'])
                ws.cell(row=row, column=3, value=entry['Section'])
                ws.cell(row=row, column=4, value=entry['Instructor'])
                ws.cell(row=row, column=5, value=entry['Room'])
                ws.cell(row=row, column=6, value=entry.get('Building', 'N/A'))
                ws.cell(row=row, column=7, value=entry['Room_Type'])
                ws.cell(row=row, column=8, value=f"{entry['Duration_Minutes']} min")

                # Apply borders
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            row += 1  # Empty row between days

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                # Skip merged cells
                if isinstance(cell, type(cell)) and hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(output_path)
        return output_path

    def _export_by_section(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each section."""

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        sections = df['Section'].unique()

        for section in sorted(sections):
            section_df = df[df['Section'] == section]

            # Create sheet
            ws = wb.create_sheet(title=str(section)[:31])  # Excel sheet name limit

            self._write_timetable_to_sheet(ws, section_df, f"Timetable for {section}")

        wb.save(output_path)
        return output_path

    def _export_by_teacher(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each teacher."""

        wb = Workbook()
        wb.remove(wb.active)

        teachers = df['Instructor'].unique()

        for teacher in sorted(teachers):
            teacher_df = df[df['Instructor'] == teacher]

            # Create sheet (sanitize name for Excel)
            safe_name = str(teacher).replace('/', '_').replace('\\', '_')[:31]
            ws = wb.create_sheet(title=safe_name)

            self._write_timetable_to_sheet(ws, teacher_df, f"Timetable for {teacher}")

        wb.save(output_path)
        return output_path

    def _export_by_room(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each room."""

        wb = Workbook()
        wb.remove(wb.active)

        rooms = df['Room'].unique()

        for room in sorted(rooms):
            room_df = df[df['Room'] == room]

            # Create sheet
            safe_name = str(room).replace('/', '_')[:31]
            ws = wb.create_sheet(title=safe_name)

            self._write_timetable_to_sheet(ws, room_df, f"Timetable for {room}")

        wb.save(output_path)
        return output_path

    def _write_timetable_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """Helper method to write timetable data to a worksheet."""

        # Title
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')

        row = 3
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        for day in days:
            day_df = df[df['Weekday'] == day].sort_values('Start_Time')

            if day_df.empty:
                continue

            # Day header
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = day
            cell.fill = self.day_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Column headers
            headers = ['Time', 'Course', 'Section', 'Room', 'Instructor', 'Duration']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.border
            row += 1

            # Data
            for _, entry in day_df.iterrows():
                ws.cell(row=row, column=1, value=f"{entry['Start_Time']}-{entry['End_Time']}")
                ws.cell(row=row, column=2, value=entry['Course_Name'])
                ws.cell(row=row, column=3, value=entry['Section'])
                ws.cell(row=row, column=4, value=entry['Room'])
                ws.cell(row=row, column=5, value=entry['Instructor'])
                ws.cell(row=row, column=6, value=f"{entry['Duration_Minutes']} min")

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            row += 1

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                if isinstance(cell, type(cell)) and hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)